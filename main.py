# Library
import pandas as pd
# time zone
from tqdm import tqdm
import numpy as np
import sys
# change width
# for error and complete
import PyQt6
from PyQt6 import QtWidgets
from openpyxl import load_workbook

# read files, if don't have file. Will say error
print("Try read excel files")
try:
    df = pd.read_excel("Article_need.xlsx")
except:
    app = QtWidgets.QApplication([])
    error_dialog = QtWidgets.QErrorMessage()
    error_dialog.setWindowTitle("Error")
    error_dialog.showMessage("Don't have Article_need.xlsx")
    app.exec()
    print("Need Article_need.xlsx")
    sys.exit()
try:
    df = pd.read_excel("Cross_have.xlsx")
except:
    app = QtWidgets.QApplication([])
    error_dialog = QtWidgets.QErrorMessage()
    error_dialog.setWindowTitle("Error")
    error_dialog.showMessage("Don't have Cross_have.xlsx")
    app.exec()
    print("Need Cross_have.xlsx")
    sys.exit()
try:
    df = pd.read_excel("DATA_JD.xlsx")
except:
    app = QtWidgets.QApplication([])
    error_dialog = QtWidgets.QErrorMessage()
    error_dialog.setWindowTitle("Error")
    error_dialog.showMessage("Don't have DATA_JD.xlsx")
    app.exec()
    print("Need DATA_JD.xlsx")
    sys.exit()

# convert from excel to data
df_jd = pd.read_excel("DATA_JD.xlsx", header = None)
articles_excel = pd.read_excel("Article_need.xlsx")
articles = articles_excel["Номер 2"] # main key

# create convenient format data
mesive = []
data = np.array(df_jd)
for work in data:
    mesive.append(list(work))

# Create key
keyboard = {}
keyboard_db = {}
resould = {}
for article in articles:
    keyboard[article] = [article]
    keyboard_db[article] = []
    resould[article] = []

# create check if don't have copy we add element
def add_in_keyboard_right(key, value):
    error = 0
    for element in keyboard[key]:
        if element == value:
            error = 1
    if error == 0:
        keyboard[key].append(value)
def add_in_keyboard_left(key, value):
    error = 0
    for element in keyboard[key]:
        if element == value:
            error = 1
    if error == 0:
        keyboard[key].insert(0,value)

# Cycle for search all data and add all data in under key elements
# {key: [here]}
temps = []
temp = []
def search_left(key, value):
    temps = []
    temps.append(value)
    while 0 == 0:
        temp = []
        for tem  in temps:
            #print(tem)
            work_value = tem
            for i, mes in enumerate(mesive):
                if mes[0] == work_value:
                    add_in_keyboard_left(key,mes[1])
                    temp.append(mes[1])
        if temp == []:
            break;
        else:
            temps = temp
def search_right(key, value):
    temps = []
    temps.append(value)
    while 0 == 0:
        temp = []
        for tem  in temps:
            #print(tem)
            work_value = tem
            for i, mes in enumerate(mesive):
                if mes[1] == work_value:
                    add_in_keyboard_left(key,mes[0])
        if temp == []:
            break;
        else:
            temps = temp
# call all def for two side
print("Choose necessary data")
for i, mes in enumerate(tqdm(mesive)):
    for art in articles:
        if mes[0] == art:
            add_in_keyboard_left(mes[0],mes[1])
            search_left(mes[0],mes[1])
            #print("Need add up ",mes)
        if mes[1] == art:
            add_in_keyboard_right(mes[1],mes[0])
            search_right(mes[1],mes[0])
            #print("Need add low ",mes)

# for DB
def add_in_keyboard_db(key, value):
    error = 0
    for element in keyboard_db[key]:
        if element == value:
            error = 1
    if error == 0:
        keyboard_db[key].append(value)

# keyboard_db element use for add data from our DB
df_have = pd.read_excel("Cross_have.xlsx")
for i, data_need in enumerate(df_have.loc[:,"Код товара"]):
    element_dd = data_need.replace("DD ","")
    #print(element_dd)
    error = 0
    for article in articles:
        if article == element_dd:
            error = 1
            #print(element_dd)
            break
    if error == 1:
        for ion,column in enumerate(df_have.columns):
            if column == "Номер перекрестной ссылки":
                code = ion
        add_in_keyboard_db(element_dd,df_have.iloc[i,ion])
# work with our Data
for article in tqdm(articles):
    for keybo in keyboard[article]:
        error = 0
        for keybo_db in keyboard_db[article]:
            if keybo == keybo_db:
                error = 1
        if error == 0:
            resould[article].append(keybo)
# Start work with view file xlsx
place_A = []
place_FH = []
for i,article in enumerate(articles):
    if len(resould[article]) != 0:
        #print(article," ",resould[article])
        for resou_art in resould[article]:
            place_A.append(article)
            place_FH.append(resou_art)

df = pd.DataFrame(np.nan,index = range(len(place_A)),columns = range(8),dtype=object)
#df.iloc[:,0] = np.array(place_A)

for i,place in enumerate(place_A):
    df.loc[i,0] = "".join(("DD ",str(place)))
for i,place in enumerate(place_FH):
    df.loc[i,5] = str(place)
    df.loc[i,7] = str(place)
df.loc[:,3] = "OE"
df.loc[:,6] = "JOHN DEERE"

#print(df)
df.to_excel("ANSWER.xlsx" ,index = False,header =False)

from openpyxl import load_workbook

# covert file in data
workbook = load_workbook("ANSWER.xlsx")
worksheet = workbook.active  # active sheet(in jupyter notebook break xlsx file(don't let change it is file)

# make size width for my data
worksheet.column_dimensions['A'].width = 20
worksheet.column_dimensions['D'].width = 5
worksheet.column_dimensions['F'].width = 15
worksheet.column_dimensions['G'].width = 12
worksheet.column_dimensions['H'].width = 15

# Save
workbook.save("ANSWER.xlsx")
# We complete all work
app = QtWidgets.QApplication([])
error_dialog = QtWidgets.QErrorMessage()
error_dialog.setWindowTitle("Completed")
error_dialog.showMessage("Created ANSWER.xlsx")
app.exec()